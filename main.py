import os
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


DB_NAME = "attendance.db"
REPORT_FOLDER = "reports"
DATE_FORMAT = "%d-%m-%Y"

os.makedirs(REPORT_FOLDER, exist_ok=True)

connection = sqlite3.connect(DB_NAME)
cursor = connection.cursor()

cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS students (
        student_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        creation_date TEXT NOT NULL
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS subjects (
        subject_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        subject_name TEXT NOT NULL,
        FOREIGN KEY (student_id) REFERENCES students (student_id)
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS attendance (
        attendance_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        subject_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        status TEXT NOT NULL,
        UNIQUE (student_id, subject_id, date),
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (subject_id) REFERENCES subjects (subject_id)
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS lecture_summary (
        summary_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        subject_id INTEGER NOT NULL,
        total_lectures INTEGER NOT NULL,
        attended_lectures INTEGER NOT NULL,
        UNIQUE (student_id, subject_id),
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (subject_id) REFERENCES subjects (subject_id)
    )
    """
)
connection.commit()


def get_students(student_id=None):
    if student_id is None:
        cursor.execute("SELECT * FROM students ORDER BY student_id")
    else:
        cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
    return cursor.fetchall()


def get_student(student_id):
    students = get_students(student_id)
    return students[0] if students else None


def get_subjects(student_id):
    cursor.execute(
        "SELECT * FROM subjects WHERE student_id = ? ORDER BY subject_id",
        (student_id,),
    )
    return cursor.fetchall()


def clean_filename(name):
    return "".join(
        character for character in name if character.isalnum() or character in (" ", "_", "-")
    ).strip().replace(" ", "_")


def read_non_negative_integer(prompt):
    try:
        value = int(input(prompt))
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        return None
    if value < 0:
        print("Please enter a non-negative number.")
        return None
    return value


def create_student():
    print("\nCreate Student")
    name = input("Enter student name: ").strip()
    if not name:
        print("Student name cannot be empty.")
        return

    creation_date = input("Enter creation date (DD-MM-YYYY): ").strip()
    try:
        datetime.strptime(creation_date, DATE_FORMAT)
    except ValueError:
        print("Invalid date format. Please use DD-MM-YYYY.")
        return

    number_of_subjects = read_non_negative_integer("Enter number of subjects: ")
    if number_of_subjects is None or number_of_subjects == 0:
        print("You need at least one subject.")
        return

    cursor.execute(
        "INSERT INTO students (name, creation_date) VALUES (?, ?)",
        (name, creation_date),
    )
    student_id = cursor.lastrowid

    for index in range(number_of_subjects):
        subject_name = input(f"Enter subject {index + 1} name: ").strip()
        if not subject_name:
            print("Subject name cannot be empty.")
            connection.rollback()
            return

        total = read_non_negative_integer("How many lectures were conducted? ")
        attended = read_non_negative_integer("How many lectures did you attend? ")
        if total is None or attended is None:
            connection.rollback()
            return
        if attended > total:
            print("Attended lectures cannot be greater than total lectures.")
            connection.rollback()
            return

        cursor.execute(
            "INSERT INTO subjects (student_id, subject_name) VALUES (?, ?)",
            (student_id, subject_name),
        )
        subject_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO lecture_summary
            (student_id, subject_id, total_lectures, attended_lectures)
            VALUES (?, ?, ?, ?)
            """,
            (student_id, subject_id, total, attended),
        )

    connection.commit()
    print(f"\nStudent created successfully! Student ID: {student_id}")
    generate_excel(student_id)


def mark_attendance():
    print("\nMark Attendance")
    students = get_students()
    if not students:
        print("No students found. Please create a student first.")
        return

    for student in students:
        print(f"{student[0]}. {student[1]}")
    try:
        student_id = int(input("\nEnter student ID: "))
    except ValueError:
        print("Invalid input. Please enter a valid student ID.")
        return

    student = get_student(student_id)
    if not student:
        print("Student not found. Please enter a valid student ID.")
        return

    date = input("Enter date (DD-MM-YYYY): ").strip()
    try:
        datetime.strptime(date, DATE_FORMAT)
    except ValueError:
        print("Invalid date format. Please use DD-MM-YYYY.")
        return

    for subject in get_subjects(student_id):
        subject_id, _, subject_name = subject
        print(f"\nSubject: {subject_name}")
        print("1. Present")
        print("2. Absent")
        print("3. No Lecture")
        choice = input("Enter your choice (1/2/3): ").strip()
        statuses = {"1": "Present", "2": "Absent", "3": "No Lecture"}
        status = statuses.get(choice)
        if status is None:
            print("Invalid choice. Skipping this subject.")
            continue

        cursor.execute(
            """
            INSERT INTO attendance (student_id, subject_id, date, status)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(student_id, subject_id, date)
            DO UPDATE SET status = excluded.status
            """,
            (student_id, subject_id, date, status),
        )

    connection.commit()
    print(f"\nAttendance saved successfully for student ID: {student_id}")
    generate_excel(student_id)


def get_attendance_data(student_id, subject_id):
    cursor.execute(
        """
        SELECT total_lectures, attended_lectures
        FROM lecture_summary
        WHERE student_id = ? AND subject_id = ?
        """,
        (student_id, subject_id),
    )
    summary = cursor.fetchone() or (0, 0)
    cursor.execute(
        """
        SELECT COUNT(*)
        FROM attendance
        WHERE student_id = ? AND subject_id = ?
        AND status IN ('Present', 'Absent')
        """,
        (student_id, subject_id),
    )
    daily_total = cursor.fetchone()[0]
    cursor.execute(
        """
        SELECT COUNT(*)
        FROM attendance
        WHERE student_id = ? AND subject_id = ? AND status = 'Present'
        """,
        (student_id, subject_id),
    )
    daily_attended = cursor.fetchone()[0]
    total = summary[0] + daily_total
    attended = summary[1] + daily_attended
    percentage = (attended / total * 100) if total else 0
    return total, attended, total - attended, percentage


def view_attendance():
    print("\nView Attendance")
    students = get_students()
    if not students:
        print("No students found. Please create a student first.")
        return

    for student in students:
        print(f"{student[0]}. {student[1]}")
    try:
        student_id = int(input("\nEnter student ID: "))
    except ValueError:
        print("Invalid input. Please enter a valid student ID.")
        return

    student = get_student(student_id)
    if not student:
        print("Student not found. Please enter a valid student ID.")
        return

    total_all = 0
    attended_all = 0
    print(f"\n{student[1]}")
    for subject in get_subjects(student_id):
        subject_id, _, subject_name = subject
        total, attended, absent, percentage = get_attendance_data(student_id, subject_id)
        total_all += total
        attended_all += attended
        print(f"\nSubject: {subject_name}")
        print(f"Total Lectures: {total}")
        print(f"Attended Lectures: {attended}")
        print(f"Absent Lectures: {absent}")
        print(f"Attendance Percentage: {percentage:.2f}%")

    overall = (attended_all / total_all * 100) if total_all else 0
    print(f"\nOverall attendance: {overall:.2f}%")


def generate_excel(student_id):
    student = get_student(student_id)
    if not student:
        print("Student not found.")
        return

    subjects = get_subjects(student_id)
    filepath = os.path.join(REPORT_FOLDER, f"{clean_filename(student[1])}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Attendance"
    sheet.append(["Date"] + [subject[2] for subject in subjects])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    cursor.execute(
        "SELECT DISTINCT date FROM attendance WHERE student_id = ? ORDER BY date",
        (student_id,),
    )
    for (date,) in cursor.fetchall():
        row = [date]
        for subject in subjects:
            cursor.execute(
                """
                SELECT status FROM attendance
                WHERE student_id = ? AND subject_id = ? AND date = ?
                """,
                (student_id, subject[0], date),
            )
            result = cursor.fetchone()
            row.append(result[0] if result else "-")
        sheet.append(row)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Student Name", student[1]])
    summary_sheet.append(["Creation Date", student[2]])
    summary_sheet.append([])
    summary_sheet.append(["Subject", "Total Lectures", "Attended", "Absent", "Attendance %"])
    for cell in summary_sheet[4]:
        cell.font = Font(bold=True)

    total_all = 0
    attended_all = 0
    for subject in subjects:
        total, attended, absent, percentage = get_attendance_data(student_id, subject[0])
        total_all += total
        attended_all += attended
        summary_sheet.append([subject[2], total, attended, absent, f"{percentage:.2f}%"])

    overall = (attended_all / total_all * 100) if total_all else 0
    summary_sheet.append([])
    summary_sheet.append(["OVERALL ATTENDANCE", "", "", "", f"{overall:.2f}%"])
    summary_sheet.cell(summary_sheet.max_row, 1).font = Font(bold=True)
    summary_sheet.cell(summary_sheet.max_row, 5).font = Font(bold=True)

    for worksheet in workbook.worksheets:
        for column in worksheet.columns:
            column_letter = get_column_letter(column[0].column)
            max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
            worksheet.column_dimensions[column_letter].width = max_length + 3

    workbook.save(filepath)
    print(f"\nExcel report created: {filepath}")


def export_excel():
    print("\n========== EXPORT EXCEL ==========")
    students = get_students()
    if not students:
        print("No students found.")
        return
    for student in students:
        print(f"{student[0]}. {student[1]}")
    try:
        student_id = int(input("\nEnter student ID: "))
    except ValueError:
        print("Invalid ID.")
        return
    generate_excel(student_id)


def main():
    while True:
        print("\n======================================")
        print("       ATTENDANCE TRACKER")
        print("======================================")
        print("1. Create new student")
        print("2. Mark daily attendance")
        print("3. View attendance")
        print("4. Generate Excel report")
        print("5. Exit")
        choice = input("Enter your choice: ").strip()

        if choice == "1":
            create_student()
        elif choice == "2":
            mark_attendance()
        elif choice == "3":
            view_attendance()
        elif choice == "4":
            export_excel()
        elif choice == "5":
            print("\nThank you for using Attendance Tracker.")
            break
        else:
            print("\nInvalid choice. Please try again.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProgram stopped.")
    finally:
        connection.close()
